import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.preprocessing import MinMaxScaler
from sklearn.model_selection import train_test_split, cross_val_score, RandomizedSearchCV
from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor
from sklearn.impute import SimpleImputer
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import Dense, Dropout
from tensorflow.keras.optimizers import Adam
from tensorflow.keras.callbacks import EarlyStopping
from tensorflow.keras.wrappers.scikit_learn import KerasRegressor
from scipy import stats
import openpyxl
from openpyxl.styles import PatternFill
import io

# Fungsi untuk memvalidasi format data
def validate_data(df):
    required_columns = ['Nama Lokasi', 'Curah Hujan', 'Suhu Udara', 'Kelembaban Udara', 'Tekanan', 'Elevasi', 'Koordinat']
    if not all(col in df.columns for col in required_columns):
        return False
    return True

# Fungsi untuk preprocessing data
def preprocess_data(data):
    # Mengubah NaN menjadi 8888
    data = data.fillna(8888)
    
    # Identifikasi kolom-kolom numerik
    numeric_columns = ['Curah Hujan', 'Suhu Udara', 'Kelembaban Udara', 'Tekanan', 'Elevasi']
    
    # Normalisasi data numerik
    scaler = MinMaxScaler()
    data[numeric_columns] = scaler.fit_transform(data[numeric_columns])
    
    return data, scaler

# Fungsi untuk membuat model ANN
def create_model(input_shape):
    model = Sequential([
        Dense(64, activation='relu', input_shape=(input_shape,)),
        Dropout(0.2),
        Dense(32, activation='relu'),
        Dropout(0.2),
        Dense(16, activation='relu'),
        Dense(1)
    ])
    model.compile(optimizer=Adam(learning_rate=0.001), loss='mse')
    return model

# Fungsi untuk evaluasi model
def evaluate_model(model, X, y):
    # Cross-validation
    cv_scores = cross_val_score(model, X, y, cv=5, scoring='neg_mean_squared_error')
    cv_rmse = np.sqrt(-cv_scores)
    
    # Prediksi
    y_pred = model.predict(X)
    
    # Metrik evaluasi
    mse = mean_squared_error(y, y_pred)
    mae = mean_absolute_error(y, y_pred)
    r2 = r2_score(y, y_pred)
    
    return {
        'RMSE': np.sqrt(mse),
        'MAE': mae,
        'R2': r2,
        'CV_RMSE_mean': cv_rmse.mean(),
        'CV_RMSE_std': cv_rmse.std()
    }

# Fungsi untuk mengukur kontribusi model
def measure_model_contributions(X_train, y_train, ann_pred, rf_pred, gb_pred):
    X_meta = np.column_stack((ann_pred, rf_pred, gb_pred))
    meta_model = GradientBoostingRegressor(n_estimators=100, random_state=42)
    meta_model.fit(X_meta, y_train)
    importances = meta_model.feature_importances_
    return {
        'ANN': importances[0],
        'Random Forest': importances[1],
        'Gradient Boosting': importances[2]
    }

# Fungsi untuk memeriksa konsistensi distribusi
def check_distribution_consistency(original_data, imputed_data):
    fig, ax = plt.subplots()
    stats.probplot(original_data, dist="norm", plot=ax)
    ax.get_lines()[0].set_markerfacecolor('blue')
    ax.get_lines()[0].set_markeredgecolor('blue')
    stats.probplot(imputed_data, dist="norm", plot=ax)
    ax.get_lines()[2].set_markerfacecolor('red')
    ax.get_lines()[2].set_markeredgecolor('red')
    ax.legend(['Original Data', 'Imputed Data'])
    ax.set_title('Q-Q Plot: Original vs Imputed Data')
    
    ks_statistic, p_value = stats.ks_2samp(original_data, imputed_data)
    
    return fig, ks_statistic, p_value

# Fungsi untuk multiple imputation
def multiple_imputation(model, X_missing, n_imputations=5):
    imputations = []
    for _ in range(n_imputations):
        imputed_values = model.predict(X_missing)
        imputations.append(imputed_values)
    return np.array(imputations)

# Fungsi untuk validasi hasil imputasi
def validate_imputation(imputed_values, column_name):
    if column_name == 'Curah Hujan':
        invalid_values = imputed_values[imputed_values < 0]
        if len(invalid_values) > 0:
            st.warning(f"Terdapat {len(invalid_values)} nilai curah hujan yang negatif setelah imputasi.")
    # Tambahkan validasi lain sesuai kebutuhan

# Fungsi untuk hyperparameter tuning
def tune_hyperparameters(X_train, y_train):
    # ANN
    def create_model(neurons=64, dropout_rate=0.2, learning_rate=0.001):
        model = Sequential([
            Dense(neurons, activation='relu', input_shape=(X_train.shape[1],)),
            Dropout(dropout_rate),
            Dense(neurons//2, activation='relu'),
            Dropout(dropout_rate),
            Dense(1)
        ])
        model.compile(optimizer=Adam(learning_rate=learning_rate), loss='mse')
        return model

    ann_model = KerasRegressor(build_fn=create_model, epochs=100, batch_size=32, verbose=0)
    ann_param_grid = {
        'neurons': [32, 64, 128],
        'dropout_rate': [0.1, 0.2, 0.3],
        'learning_rate': [0.001, 0.01, 0.1]
    }
    ann_random = RandomizedSearchCV(estimator=ann_model, param_distributions=ann_param_grid, n_iter=10, cv=3, random_state=42)
    ann_random.fit(X_train, y_train)

    # Random Forest
    rf_param_grid = {
        'n_estimators': [100, 200, 300],
        'max_depth': [10, 20, 30, None],
        'min_samples_split': [2, 5, 10],
        'min_samples_leaf': [1, 2, 4]
    }
    rf_random = RandomizedSearchCV(estimator=RandomForestRegressor(random_state=42), param_distributions=rf_param_grid, n_iter=10, cv=3, random_state=42)
    rf_random.fit(X_train, y_train)

    return ann_random.best_estimator_, rf_random.best_estimator_

# Fungsi utama untuk training dan imputasi
def train_and_impute(data, target_column, scaler):
    X = data.drop(columns=[target_column, 'Nama Lokasi', 'Koordinat'])
    y = data[target_column]
    
    X_complete = X[y != 8888]
    y_complete = y[y != 8888]
    X_missing = X[y == 8888]
    
    X_train, X_test, y_train, y_test = train_test_split(X_complete, y_complete, test_size=0.2, random_state=42)
    
    # Hyperparameter tuning
    ann_model, rf_model = tune_hyperparameters(X_train, y_train)
    
    # Train models
    ann_model.fit(X_train, y_train)
    rf_model.fit(X_train, y_train)
    gb_model = GradientBoostingRegressor(n_estimators=100, random_state=42)
    gb_model.fit(X_train, y_train)
    
    # Predictions
    ann_train_pred = ann_model.predict(X_train)
    rf_train_pred = rf_model.predict(X_train)
    gb_train_pred = gb_model.predict(X_train)
    
    ann_test_pred = ann_model.predict(X_test)
    rf_test_pred = rf_model.predict(X_test)
    gb_test_pred = gb_model.predict(X_test)
    
    # Model contributions
    model_contributions = measure_model_contributions(X_train, y_train, ann_train_pred, rf_train_pred, gb_train_pred)
    
    # Combine predictions
    final_train_pred = (ann_train_pred + rf_train_pred + gb_train_pred) / 3
    final_test_pred = (ann_test_pred + rf_test_pred + gb_test_pred) / 3
    
    # Evaluate models
    ann_metrics = evaluate_model(ann_model, X_test, y_test)
    rf_metrics = evaluate_model(rf_model, X_test, y_test)
    gb_metrics = evaluate_model(gb_model, X_test, y_test)
    final_metrics = evaluate_model(lambda X: (ann_model.predict(X) + rf_model.predict(X) + gb_model.predict(X)) / 3, X_test, y_test)
    
    # Multiple imputation
    ann_imputations = multiple_imputation(ann_model, X_missing)
    rf_imputations = multiple_imputation(rf_model, X_missing)
    gb_imputations = multiple_imputation(gb_model, X_missing)
    
    final_imputations = (ann_imputations + rf_imputations + gb_imputations) / 3
    y_imputed = final_imputations.mean(axis=0)
    imputation_std = final_imputations.std(axis=0)
    
    # Inverse transform imputed values
    y_imputed_original_scale = scaler.inverse_transform(np.column_stack((y_imputed, X_missing)))[:, 0]
    
    # Validate imputation
    validate_imputation(y_imputed_original_scale, target_column)
    
    return y_imputed_original_scale, imputation_std, final_metrics, model_contributions, ann_metrics, rf_metrics, gb_metrics, final_train_pred, y_train, final_test_pred, y_test

# Fungsi untuk menyimpan hasil ke Excel
def save_to_excel(data, imputed_values, imputation_std, target_column):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        data.to_excel(writer, index=False, sheet_name='Sheet1')
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        fill_imputed = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        fill_original = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        
        for i, row in data.iterrows():
            cell = worksheet.cell(row=i+2, column=data.columns.get_loc(target_column)+1)
            if row[target_column] == 8888:
                cell.value = imputed_values[data[data[target_column] == 8888].index.get_loc(i)]
                cell.fill = fill_imputed
                # Add imputation standard deviation
                std_cell = worksheet.cell(row=i+2, column=data.columns.get_loc(target_column)+2)
                std_cell.value = imputation_std[data[data[target_column] == 8888].index.get_loc(i)]
            else:
                cell.fill = fill_original
    
    output.seek(0)
    return output

# Fungsi utama Streamlit
def main():
    st.set_page_config(page_title="Dashboard Imputasi Data Curah Hujan", layout="wide")
    
    st.title("Dashboard Imputasi Data Curah Hujan")
    
    uploaded_file = st.sidebar.file_uploader("Unggah file Excel", type="xlsx")
    
    if uploaded_file is not None:
        data = pd.read_excel(uploaded_file)
        if not validate_data(data):
            st.error("Format data tidak sesuai. Pastikan file Excel memiliki kolom: Nama Lokasi, Curah Hujan, Suhu Udara, Kelembaban Udara, Tekanan, Elevasi, Koordinat")
            return
    else:
        st.warning("Silakan unggah file Excel dengan data curah hujan")
        return
    
    target_column = st.sidebar.selectbox("Pilih kolom target untuk imputasi", ['Curah Hujan', 'Suhu Udara', 'Kelembaban Udara', 'Tekanan'])
    
    data_processed, scaler = preprocess_data(data)
    
    if st.button("Lakukan Imputasi"):
        with st.spinner("Sedang melakukan imputasi..."):
            imputed_values, imputation_std, final_metrics, model_contributions, ann_metrics, rf_metrics, gb_metrics, final_train_pred, y_train, final_test_pred, y_test = train_and_impute(data_processed, target_column, scaler)
        
        st.success("Imputasi selesai!")
        
        # Tampilkan metrik
        st.subheader("Metrik Evaluasi")
        col1, col2, col3 = st.columns(3)
        col1.metric("RMSE", f"{final_metrics['RMSE']:.4f}")
        col2.metric("MAE", f"{final_metrics['MAE']:.4f}")
        col3.metric("R2", f"{final_metrics['R2']:.4f}")
        
        # Perbandingan model
        st.subheader("Perbandingan Model")
        comparison_df = pd.DataFrame({
            'ANN': ann_metrics,
            'Random Forest': rf_metrics,
            'Gradient Boosting': gb_metrics,
            'Ensemble': final_metrics
        }).T
        st.dataframe(comparison_df)
        
        # Model contributions
        st.subheader("Kontribusi Model")
        st.write(model_contributions)
        
       # Visualisasi: scatter plot prediksi vs aktual
        st.subheader("Prediksi vs Aktual")
        fig, ax = plt.subplots()
        ax.scatter(y_test, final_test_pred, alpha=0.5)
        ax.plot([y_test.min(), y_test.max()], [y_test.min(), y_test.max()], 'r--', lw=2)
        ax.set_xlabel("Nilai Aktual")
        ax.set_ylabel("Nilai Prediksi")
        ax.set_title("Scatter Plot: Prediksi vs Aktual")
        st.pyplot(fig)

        # Visualisasi: box plot perbandingan distribusi sebelum dan sesudah imputasi
        st.subheader("Distribusi Sebelum dan Sesudah Imputasi")
        fig, ax = plt.subplots()
        data_before = data[data[target_column] != 8888][target_column]
        data_after = data.copy()
        data_after.loc[data_after[target_column] == 8888, target_column] = imputed_values
        sns.boxplot(data=[data_before, data_after[target_column]], ax=ax)
        ax.set_xticklabels(['Sebelum Imputasi', 'Setelah Imputasi'])
        ax.set_title(f"Distribusi {target_column}")
        st.pyplot(fig)

        # Cek konsistensi distribusi
        original_data = data[data[target_column] != 8888][target_column]
        fig, ks_statistic, p_value = check_distribution_consistency(original_data, imputed_values)
        st.subheader("Konsistensi Distribusi")
        st.pyplot(fig)
        st.write(f"Kolmogorov-Smirnov statistic: {ks_statistic:.4f}")
        st.write(f"p-value: {p_value:.4f}")

        # Tampilkan data gabungan
        st.subheader("Data Gabungan (Asli + Imputasi)")
        data_combined = data.copy()
        data_combined.loc[data_combined[target_column] == 8888, target_column] = imputed_values
        st.dataframe(data_combined)

        # Analisis sensitivitas
        st.subheader("Analisis Sensitivitas")
        original_mean = data[data[target_column] != 8888][target_column].mean()
        imputed_mean = data_combined[target_column].mean()
        st.write(f"Mean sebelum imputasi: {original_mean:.2f}")
        st.write(f"Mean setelah imputasi: {imputed_mean:.2f}")
        st.write(f"Perubahan mean: {(imputed_mean - original_mean) / original_mean * 100:.2f}%")

        # Tampilkan ketidakpastian imputasi
        st.subheader("Ketidakpastian Imputasi")
        st.write(f"Rata-rata standar deviasi imputasi: {imputation_std.mean():.4f}")
        
        # Visualisasi ketidakpastian imputasi
        fig, ax = plt.subplots()
        ax.hist(imputation_std, bins=20)
        ax.set_xlabel("Standar Deviasi")
        ax.set_ylabel("Frekuensi")
        ax.set_title("Distribusi Ketidakpastian Imputasi")
        st.pyplot(fig)

        # Simpan hasil ke Excel
        output = save_to_excel(data_combined, imputed_values, imputation_std, target_column)
        st.download_button(
            label="Unduh Hasil Imputasi",
            data=output,
            file_name="hasil_imputasi.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # Tampilkan peta lokasi
    st.subheader("Peta Lokasi")
    if 'Koordinat' in data.columns:
        koordinat = data['Koordinat'].str.split(',', expand=True).astype(float)
        data['Latitude'] = koordinat[0]
        data['Longitude'] = koordinat[1]
        st.map(data)
    else:
        st.warning("Data koordinat tidak tersedia")

    # Tampilkan data asli
    st.subheader("Data Asli")
    st.dataframe(data)

if __name__ == "__main__":
    main()
